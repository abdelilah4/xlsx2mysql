try:
    from openpyxl import load_workbook
except:
    print 'module openpyxl not installed'
def Readxlsx(target,page):
    wb = load_workbook(filename=target, read_only=True)
    ws = wb[page]
    col = ws.max_column
    j=0
    k=0
    result = list()
    minlist = list()
    for row in ws.rows:
        for cell in row:
            minlist.append(cell.value)
            j = j+1
            if j==(k+col):
                result.append(minlist)
                minlist = list()
                k=k+col
    return result
def infoxlsx(target):
    wb = load_workbook(filename=target, read_only=True)
    result = list()
    tis = wb.worksheets
    for i in tis:
        result.append(i.title)
    return result
